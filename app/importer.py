"""One-time, provenance-conscious import of the supplied CSV and workbook."""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.database import Database

STANDARD_CAGE_ID = re.compile(r"^CC\d{8}$", flags=re.IGNORECASE)


@dataclass(slots=True)
class ImportReport:
    seeded: bool = False
    cages: int = 0
    animals: int = 0
    enriched_animals: int = 0
    surgeries: int = 0
    warnings: list[str] = field(default_factory=list)


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _date_text(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _text(value)
    if not text:
        return None
    for pattern in (
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%m/%d/%Y %I:%M %p",
        "%Y-%m-%d %H:%M:%S",
    ):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            continue
    return None


def _integer(value: Any) -> int:
    text = _text(value)
    if not text:
        return 0
    try:
        return max(0, int(float(text)))
    except ValueError:
        return 0


def _status(value: Any) -> str:
    normalized = _text(value).casefold()
    if normalized == "active":
        return "active"
    if normalized in {"on order", "on_order", "ordered"}:
        return "on_order"
    return "inactive"


def _append_note(current: str | None, addition: str | None) -> str | None:
    existing = _text(current)
    incoming = _text(addition)
    if not incoming or incoming in existing:
        return existing or None
    return f"{existing}\n{incoming}".strip() if existing else incoming


def _genotype(value: Any) -> str | None:
    raw = _text(value)
    if not raw:
        return None
    return "WT" if raw.casefold() == "wt" else raw


def _canonical_surgery_type(value: Any) -> str:
    raw = _text(value)
    folded = " ".join(raw.casefold().split())
    if folded in {
        "headplate + probe implant",
        "headplate+probe implant",
        "headplate + probeimplant",
        "headplate+probeimplant",
        "headplate and probe implant",
        "headplate + npx",
        "headplate and npx",
        "headplate + neuropixel",
        "headplate and neuropixel",
    }:
        return "Headplate + probe implant"
    if folded in {"headplate", "head plate"}:
        return "Headplate"
    if folded in {
        "probe",
        "probe implant",
        "npx",
        "npx implant",
        "neuropixel",
        "neuropixel implant",
    }:
        return "Probe implant"
    return raw or "Surgery"


def _sheet_records(workbook_path: Path) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    sheets: dict[str, list[dict[str, Any]]] = {}
    try:
        for worksheet in workbook.worksheets:
            rows = list(worksheet.iter_rows(values_only=True))
            if not rows:
                sheets[worksheet.title] = []
                continue
            headers = [_text(value) for value in rows[0]]
            records: list[dict[str, Any]] = []
            for row_number, values in enumerate(rows[1:], start=2):
                record = {
                    header: values[index] if index < len(values) else None
                    for index, header in enumerate(headers)
                    if header
                }
                substantive = any(
                    value not in (None, "")
                    and not (isinstance(value, str) and value.startswith("="))
                    for value in values
                )
                if substantive:
                    record["_row"] = row_number
                    records.append(record)
            sheets[worksheet.title] = records
    finally:
        workbook.close()
    return sheets


def _is_single_character_edit(left: str, right: str) -> bool:
    if abs(len(left) - len(right)) > 1:
        return False
    if len(left) == len(right):
        return sum(a != b for a, b in zip(left, right, strict=True)) == 1
    shorter, longer = (left, right) if len(left) < len(right) else (right, left)
    short_index = 0
    long_index = 0
    edits = 0
    while short_index < len(shorter) and long_index < len(longer):
        if shorter[short_index] == longer[long_index]:
            short_index += 1
            long_index += 1
            continue
        edits += 1
        long_index += 1
        if edits > 1:
            return False
    return True


def _resolve_cage_id(value: Any, cage_ids: dict[str, int]) -> str | None:
    candidate = _text(value).upper()
    if not candidate:
        return None
    if candidate in cage_ids:
        return candidate
    matches = [cage_key for cage_key in cage_ids if _is_single_character_edit(candidate, cage_key)]
    return matches[0] if len(matches) == 1 else None


def _match_cage_id(record: dict[str, Any], cage_ids: dict[str, int]) -> str | None:
    direct = _text(record.get("Cage Card #"))
    if direct:
        return _resolve_cage_id(direct, cage_ids)
    return _resolve_cage_id(record.get("Seperated From"), cage_ids)


def _apply_split_lineage(
    database: Database,
    sheets: dict[str, list[dict[str, Any]]],
    cage_ids: dict[str, int],
    report: ImportReport,
) -> None:
    applied: set[tuple[str, str]] = set()
    for sheet_name, records in sheets.items():
        for record in records:
            direct = _resolve_cage_id(record.get("Cage Card #"), cage_ids)
            source = _resolve_cage_id(record.get("Seperated From"), cage_ids)
            if direct is None or source is None or direct == source:
                continue
            pair = (direct, source)
            if pair in applied:
                continue
            try:
                database.set_imported_split_lineage(cage_ids[direct], cage_ids[source])
            except ValueError as exc:
                report.warnings.append(f"{sheet_name} row {record['_row']} lineage skipped: {exc}")
            else:
                applied.add(pair)


def _apply_main_sheet(
    database: Database,
    records: list[dict[str, Any]],
    cage_ids: dict[str, int],
    report: ImportReport,
) -> None:
    for record in records:
        cage_key = _match_cage_id(record, cage_ids)
        if cage_key is None:
            report.warnings.append(f"Main row {record['_row']} has no matching CSV cage.")
            continue
        cage_id = cage_ids[cage_key]
        animals = database.list_animals(cage_id, include_inactive=True)
        male_count = _integer(record.get("male"))
        female_count = _integer(record.get("female"))
        dob = _date_text(record.get("DOB"))
        genotype = _genotype(record.get("Strain"))
        requested = male_count + female_count
        if requested and requested != len(animals):
            report.warnings.append(
                f"{cage_key} Excel sex total {requested} differs from authoritative "
                f"CSV count {len(animals)}."
            )
        for index, animal in enumerate(animals):
            sex = animal["sex"]
            if index < male_count:
                sex = "M"
            elif index < male_count + female_count:
                sex = "F"
            database.update_animal(
                animal["id"],
                sex=sex,
                dob=dob or animal.get("dob"),
                genotype=genotype or animal.get("genotype"),
                note=animal.get("note"),
                legacy_id=animal.get("legacy_id"),
            )
        if animals:
            target = animals[0]
            legacy_id = _text(record.get("mouse#")) or target.get("legacy_id")
            if legacy_id:
                database.update_animal(target["id"], legacy_id=legacy_id)
            operator = _text(record.get("User")) or "Unknown"
            for date_field, type_field in (
                ("Date surgery #1", "surg#1 procedure"),
                ("Date surgery #2", "surg#2 procedure"),
            ):
                surgery_date = _date_text(record.get(date_field))
                if not surgery_date:
                    continue
                try:
                    created = database.add_surgery(
                        target["id"],
                        surgery_date=surgery_date,
                        surgery_time=None,
                        operator=operator,
                        surgery_type=_canonical_surgery_type(record.get(type_field)),
                    )
                except ValueError as exc:
                    report.warnings.append(f"Main row {record['_row']} surgery skipped: {exc}")
                else:
                    if created:
                        report.surgeries += 1
        cage = database.get_cage(cage_id)
        if cage is not None:
            database.update_cage(
                cage_id,
                note=_append_note(cage.get("note"), _text(record.get("Note")) or None),
            )


def _select_individual(
    animals: list[dict[str, Any]],
    claimed: set[int],
    *,
    legacy_id: str | None,
    sex: str,
    dob: str | None,
) -> dict[str, Any] | None:
    if legacy_id:
        for animal in animals:
            if _text(animal.get("legacy_id")) == legacy_id:
                return animal
    available = [animal for animal in animals if animal["id"] not in claimed]
    for animal in available:
        if sex != "U" and animal.get("sex") not in {sex, "U"}:
            continue
        if dob and animal.get("dob") not in {None, "", dob}:
            continue
        return animal
    return available[0] if available else None


def _apply_individual_sheets(
    database: Database,
    sheets: dict[str, list[dict[str, Any]]],
    cage_ids: dict[str, int],
    report: ImportReport,
) -> None:
    claimed: set[int] = set()
    for sheet_name, records in sheets.items():
        if sheet_name == "Main":
            continue
        for record in records:
            cage_key = _match_cage_id(record, cage_ids)
            if cage_key is None:
                report.warnings.append(
                    f"{sheet_name} row {record['_row']} has no matching CSV cage."
                )
                continue
            cage_id = cage_ids[cage_key]
            legacy_id = _text(record.get("mouse#")) or None
            male = _integer(record.get("male"))
            female = _integer(record.get("female"))
            sex = "M" if male else "F" if female else "U"
            dob = _date_text(record.get("DOB"))
            genotype = _genotype(record.get("Strain"))
            animals = database.list_animals(cage_id, include_inactive=True)
            animal = _select_individual(
                animals,
                claimed,
                legacy_id=legacy_id,
                sex=sex,
                dob=dob,
            )
            if animal is None:
                report.warnings.append(
                    f"{sheet_name} row {record['_row']} has no unclaimed CSV mouse slot."
                )
                continue
            claimed.add(animal["id"])
            database.update_animal(
                animal["id"],
                legacy_id=legacy_id or animal.get("legacy_id"),
                sex=sex if sex != "U" else animal.get("sex", "U"),
                dob=dob or animal.get("dob"),
                genotype=genotype or animal.get("genotype"),
                note=_append_note(animal.get("note"), _text(record.get("Note")) or None),
            )
            report.enriched_animals += 1

            operator = _text(record.get("User")) or sheet_name
            for date_field, type_field in (
                ("Date surgery #1", "surg#1 procedure"),
                ("Date surgery #2", "surg#2 procedure"),
            ):
                surgery_date = _date_text(record.get(date_field))
                if not surgery_date:
                    continue
                try:
                    created = database.add_surgery(
                        animal["id"],
                        surgery_date=surgery_date,
                        surgery_time=None,
                        operator=operator,
                        surgery_type=_canonical_surgery_type(record.get(type_field)),
                    )
                except ValueError as exc:
                    report.warnings.append(
                        f"{sheet_name} row {record['_row']} surgery skipped: {exc}"
                    )
                else:
                    if created:
                        report.surgeries += 1


def seed_if_empty(
    database: Database,
    *,
    csv_path: Path,
    xlsx_path: Path,
) -> ImportReport:
    """Seed an empty database, using CSV cage counts as the source of truth."""

    report = ImportReport()
    if not database.is_empty():
        return report
    if not csv_path.is_file():
        report.warnings.append(f"CSV source not found: {csv_path}")
        return report

    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    rows.sort(key=lambda row: 0 if _status(row.get("Status")) == "active" else 1)

    cage_ids: dict[str, int] = {}
    for row_number, row in enumerate(rows, start=2):
        cage_card_id = _text(row.get("Cage Card ID")).upper()
        if not cage_card_id:
            report.warnings.append(f"CSV row {row_number} has no Cage Card ID.")
            continue
        if not STANDARD_CAGE_ID.fullmatch(cage_card_id):
            report.warnings.append(f"CSV row {row_number} has a nonstandard Cage Card ID.")
        count = _integer(row.get("# Animals"))
        cage_id = database.create_cage(
            cage_card_id=cage_card_id,
            status=_status(row.get("Status")),
            animal_count=count,
            sex="U",
            dob=None,
            genotype=None,
            room=_text(row.get("Room")) or None,
            protocol=_text(row.get("Protocol")) or None,
            note=None,
            creation_type="import",
            on_census_date=_date_text(row.get("On Census Date")),
            off_census_date=_date_text(row.get("Off Census Date")),
        )
        cage_ids[cage_card_id] = cage_id
        report.cages += 1
        report.animals += count

    if xlsx_path.is_file():
        sheets = _sheet_records(xlsx_path)
        _apply_split_lineage(database, sheets, cage_ids, report)
        _apply_main_sheet(database, sheets.get("Main", []), cage_ids, report)
        _apply_individual_sheets(database, sheets, cage_ids, report)
    else:
        report.warnings.append(f"Excel source not found: {xlsx_path}")

    report.seeded = True
    return report
